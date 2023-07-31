import openpyxl
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import warnings
import os

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=FutureWarning)


# tạo exel với số lượng sheet cho trước
def create_excel(filename, sheet_names):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name in sheet_names:
        workbook.create_sheet(title=sheet_name)
    workbook.save(filename)
    print(f"--- Tạo tệp {filename} thành công với {len(sheet_names)} sheet!")


# format sheet đầu tiên file excel
def format_fisrt_sheet(file_path, sheet_name, start_cell, end_cell, color):
    # print("================", file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell_range = sheet[start_cell:end_cell]
    alignment = Alignment(horizontal="left", vertical="center")
    for row in cell_range:
        for cell in row:
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.alignment = alignment
            cell.font = Font(name="ＭＳ Ｐゴシック", size=24, bold=False)
    sheet.row_dimensions[1].height = 60
    # print("===========", sheet.row_dimensions[1].height)
    workbook.save(file_path)


# lấy data từ test list
def get_data_from_testlist(source_file, destination_file):
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet_name = source_workbook.sheetnames[0]
    source_sheet = source_workbook[source_sheet_name]
    destination_workbook = openpyxl.Workbook()
    destination_sheet = destination_workbook.active
    destination_sheet.append([""])
    test_case = 1
    arr_skip = ["チェックリスト", None, "No"]
    for row in source_sheet.iter_rows(values_only=True):
        # print(row)
        value_in_column_a = row[0]  # Get the value in column A
        value_in_column_b = row[1]
        # print("aaaaa", value_in_column_a)
        if value_in_column_a != None and value_in_column_a != "No":
            if arr_skip[0] not in str(value_in_column_a):  # Get the value in column B
                destination_sheet.append([value_in_column_b])
                # print("jghjhgj", value_in_column_a)
                print("Get thành công test care:", test_case, value_in_column_b)
                test_case += 1
    destination_workbook.save(destination_file)
    print("--- Tạo thành công file", destination_file)
    return test_case


# format file test evidence với điều kiện phải định dạng sheet đầu tiên của file
def copy_format_sheet(
    file_name, source_sheet, destination_sheet, source_range, destination_range
):
    workbook = openpyxl.load_workbook(file_name)
    source_sheet = workbook[source_sheet]
    destination_sheet = workbook[destination_sheet]
    source_cells = source_sheet[source_range]
    destination_cells = destination_sheet[destination_range]
    for src_row, dest_row in zip(source_cells, destination_cells):
        for src_cell, dest_cell in zip(src_row, dest_row):
            dest_cell.number_format = src_cell.number_format
            dest_cell.font = src_cell.font.copy()
            dest_cell.alignment = src_cell.alignment.copy()
            dest_cell.border = src_cell.border.copy()
            dest_cell.fill = src_cell.fill.copy()
    destination_sheet.row_dimensions[1].height = 60
    # print(destination_sheet.row_dimensions[1].height)
    workbook.save(file_name)


def format_full_sheet(file_name, sl_sheet):
    source_sheet = "1"
    source_range = "A1:Z1"
    destination_range = "A1:Z1"
    sheet_name = "1"
    format_fisrt_sheet(file_name, sheet_name, "A1", "Z1", "3F9FFF")
    destination_sheet_name_arr = [str(i + 1) for i in range(1, sl_sheet)]
    for destination_sheet_name in destination_sheet_name_arr:
        # print(destination_sheet_name)
        copy_format_sheet(
            file_name,
            source_sheet,
            destination_sheet_name,
            source_range,
            destination_range,
        )
    print("--- Format file", file_name, "thành công")


# import data từ data test list
def import_data(
    input_file,
    output_file,
    source_sheet,
    source_cell_arr,
    target_sheets_arr,
    target_cell,
):
    df = pd.read_excel(input_file, sheet_name=source_sheet)
    writer = pd.ExcelWriter(output_file, engine="openpyxl", mode="a")
    writer.book = load_workbook(output_file)
    count_sheet = 0
    for source_cell in source_cell_arr:
        source_value = df.iloc[source_cell[0], source_cell[1]]  # Lấy giá trị từ ô nguồn
        # print("import data:", count_sheet + 1, source_value)
        sheet_name = target_sheets_arr[count_sheet]
        writer.book[sheet_name][target_cell].value = source_value
        count_sheet += 1
    writer.save()
    writer.close()
    print("--- Hoàn thành tạo file test evidence:", output_file)


def name_sheet(sl_sheet):
    name_sheets_arr = []
    for i in range(sl_sheet):
        name_sheets_arr.append(str(i + 1))
    return name_sheets_arr


def count_sheets_xlsx(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        return len(workbook.sheetnames)
    except Exception as e:
        print("Error:", e)
        return 0


def run(filename):
    try:
        file_list = os.listdir("processed")
        for filename in file_list:
            file_path = os.path.join("processed", filename)
            os.remove(file_path)
        name_file = filename

        test_list = "uploads/" + name_file
        test_evidence = "processed/evidence_" + name_file
        draw_data = "processed/draw_" + name_file
        # sl_sheet = count_sheets_xlsx(test_list)

        sl_sheet = get_data_from_testlist(test_list, draw_data) - 1
        arr_sheet = name_sheet(sl_sheet)
        source_cell_arr = []
        target_cell = "A1"
        for i in range(sl_sheet):
            source_cell_arr.append((i, 0))
        # print(source_cell_arr)
        create_excel(test_evidence, arr_sheet)
        format_full_sheet(test_evidence, sl_sheet)
        sheet_draw_data = "Sheet"
        import_data(
            draw_data,
            test_evidence,
            sheet_draw_data,
            source_cell_arr,
            arr_sheet,
            target_cell,
        )
    except Exception as e:
        print("Error:", e)
